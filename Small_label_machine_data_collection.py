import time,threading,queue,os,datetime,snap7
from openpyxl import Workbook,load_workbook
from snap7.util import get_bool,set_bool
# import datetime

setting_file=open('small_label_machine_1_settings.txt','r')
print("MT: I am a Small label data backup maker")

filedic={}
for line in setting_file:
    file_data=line.strip().split('===')
    a=file_data[0]
    b=file_data[1]
    filedic[a]=b
setting_file.close()

shiftA_start=filedic.pop('shiftA_start_time')
shiftB_start=filedic.pop('shiftB_start_time')
shiftC_start=filedic.pop('shiftC_start_time')
filename_of_excel=filedic.pop('filename_of_excel_sheet')
ipaddress_of_plc=filedic.pop('ipaddress_of_plc')
db_number=int(filedic.pop('data_block_number_of_plc'))
rack_number=int(filedic.pop('rack_number_of_plc'))
slot_number=int(filedic.pop('slot_number_of_plc'))
plc_db_read_delay=int(filedic.pop('plc_db_read_delay_in_milliseconds'))/1000

A=list(map(int,shiftA_start.strip().split(":")))
B=list(map(int,shiftB_start.strip().split(":")))
C=list(map(int,shiftC_start.strip().split(":")))

excel_queue=queue.Queue()

def get_shift(ct):
    startA=datetime.time(A[0],A[1],A[2])
    startB=datetime.time(B[0],B[1],B[2])
    startC=datetime.time(C[0],C[1],C[2])
    if startA<=ct<startB:
        return 'A'
    elif startB<=ct<startC:
        return 'B'
    else:
        return 'C'

def snap7_thread():
    connection_flag=False
    while True:
        try:
            if not connection_flag:
                client=snap7.client.Client()
                client.connect(ipaddress_of_plc,rack_number,slot_number)
                connection_flag=True
            prim_data=client.db_read(db_number,0,1)
            #print(prim_data)
            register_flag=get_bool(prim_data,0,0)
            if register_flag:
                data_byte=client.db_read(db_number,2,512)
                sc_barcode_data=data_byte[2:2+data_byte[1]].decode()
                printer_data=data_byte[258:258+data_byte[257]].decode()
                print("SC:",sc_barcode_data," PR:",printer_data)
                set_bool(prim_data,0,0,0)
                excel_queue.put([sc_barcode_data,printer_data])
                client.db_write(db_number,0,prim_data)
            time.sleep(plc_db_read_delay)
        except Exception as e:
            if str(e)=="b' TCP : Unreachable peer'":
                print(f' Unable to connect PLC')
                connection_flag=False
            if str(e)=="b' ISO : An error occurred during send TCP : Connection reset by peer'":
                print(f' Unable to connect PLC')
                connection_flag=False
            if str(e)=="b' ISO : An error occurred during recv TCP : Connection timed out'":
                print(f' Unable to connect PLC')
                connection_flag=False
            else:
                print(e)
            time.sleep(2)

def move_excel():
    while True:
        try:
            s_data,p_data=excel_queue.get()
            # data_list=[]
            # print( len(s_data))
            t=datetime.datetime.now()
            current_time=datetime.time(t.hour,t.minute,t.second)
            if current_time<datetime.time(A[0],A[1],A[2]):
                dp=1
            else:
                dp=0
            date=t-datetime.timedelta(days=dp)
            sheet_name=(date.strftime("%b%Y")).upper()
            date=date.strftime("%d-%m-%Y")
            if not os.path.isfile(filename_of_excel):
                wb=Workbook()
            else:
                wb=load_workbook(filename_of_excel)
            if not sheet_name in wb.sheetnames:
                wb.create_sheet(sheet_name)
                ws=wb[sheet_name]
            ws=wb[sheet_name]        
            excel_dic={}
            excel_dic["DATE"]=date
            excel_dic["SHIFT"]=get_shift(current_time)
            excel_dic["SC_DATA"]=s_data
            excel_dic["PR_DATA"]=p_data
            excel_dic["TIME"]=time.strftime("%d-%m-%Y_%I.%M.%S_%p")
            xl_headers=[]
            for i in ws[1]:
                xl_headers.append(i.value)
            mc=ws.max_column
            mr=ws.max_row
            for i in excel_dic:
                if i not in xl_headers:
                    ws.cell(1,mc+1).value=i
                    xl_headers.append(i)
                    mc=ws.max_column
            for index_i,i in enumerate(xl_headers):
                for j in excel_dic:
                    if i==j:
                        ws.cell(mr+1,index_i+1).value=excel_dic[i]
            while True:
                try:
                    wb.save(filename_of_excel)
                    print(f"EXP: Part no Data saved successfully {time.strftime('%d-%m-%Y_%I.%M.%S_%p')} {line}")
                except Exception as e:
                    print(f"EXP: Data not saved, Close the excel file({filename_of_excel}) if it is opened. Retrying to save...\n Error: {e}")
                    time.sleep(3)
                    continue
                break
            wb.close()
        except Exception as e:
            print(e)


snap7_th=threading.Thread(target=snap7_thread,daemon=True)
snap7_th.start()
print("MT: Snap7_thread thread started")
move_ex = threading.Thread(target=move_excel,daemon=True)
move_ex.start()
print("MT: Move_excel thread started")

while True:
    time.sleep(60)